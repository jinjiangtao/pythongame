from model.bill_model import BillModel
from model.category_model import CategoryModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os


class BillController:
    def __init__(self, view, user):
        self.view = view
        self.user = user
        self.bill_model = BillModel()
        self.category_model = CategoryModel()
        self.view.set_add_command(self.handle_add)
        self.view.set_export_command(self.handle_export)
        self.view.set_search_command(self.handle_search)
        self.view.set_statistics_command(self.handle_statistics)
        self.view.set_prev_page_command(self.handle_prev_page)
        self.view.set_next_page_command(self.handle_next_page)
        self.view.set_edit_callback(self.handle_edit)
        self.view.set_delete_callback(self.handle_delete)
        self.load_categories()
        self.load_bills()

    def load_categories(self):
        categories = self.category_model.get_categories(self.user["id"])
        self.view.set_categories(categories)

    def load_bills(self):
        params = self.view.get_filter_params()
        category_id = None
        if params["category_id"]:
            try:
                category_id = int(params["category_id"].split(":")[0])
            except (ValueError, IndexError):
                category_id = None

        current_page = self.view.get_current_page()
        page_size = self.view.get_page_size()

        bills = self.bill_model.get_bills(
            self.user["id"],
            type_=params["type"],
            category_id=category_id,
            start_date=params["start_date"],
            end_date=params["end_date"],
            page=current_page,
            page_size=page_size
        )

        total_count = self.bill_model.get_bills_count(
            self.user["id"],
            type_=params["type"],
            category_id=category_id,
            start_date=params["start_date"],
            end_date=params["end_date"]
        )

        total_pages = (total_count + page_size - 1) // page_size
        self.view.update_pagination(current_page, max(total_pages, 1), total_count)
        self.view.display_bills(bills)
        self.update_summary()

    def handle_search(self):
        self.view.reset_page()
        self.load_bills()

    def handle_prev_page(self):
        current_page = self.view.get_current_page()
        if current_page > 1:
            self.view.update_pagination(current_page - 1, self.view.total_pages, self.view.total_count)
            self.load_bills()

    def handle_next_page(self):
        current_page = self.view.get_current_page()
        if current_page < self.view.total_pages:
            self.view.update_pagination(current_page + 1, self.view.total_pages, self.view.total_count)
            self.load_bills()

    def update_summary(self):
        params = self.view.get_filter_params()
        category_id = None
        if params["category_id"]:
            try:
                category_id = int(params["category_id"].split(":")[0])
            except (ValueError, IndexError):
                category_id = None

        summary = self.bill_model.get_summary(
            self.user["id"],
            type_=params["type"],
            category_id=category_id,
            start_date=params["start_date"],
            end_date=params["end_date"]
        )
        self.view.update_summary(summary)

    def handle_add(self):
        categories = self.category_model.get_categories(self.user["id"])
        self.view.show_add_dialog(categories, self.save_bill)

    def save_bill(self, type_, category_id, amount_val, remark, payment_method, date, bill_id=None):
        try:
            amount = float(amount_val)
        except ValueError:
            self.view.show_message("金额必须是数字", is_error=True)
            return

        if bill_id:
            success, message = self.bill_model.update_bill(
                bill_id, self.user["id"], type_, category_id, amount, remark, payment_method, date
            )
        else:
            success, message = self.bill_model.add_bill(
                self.user["id"], type_, category_id, amount, remark, payment_method, date
            )
        
        if success:
            self.view.show_message(message)
            self.load_bills()
        else:
            self.view.show_message(message, is_error=True)

    def handle_edit(self, bill_id):
        bill = self.bill_model.get_bill_by_id(bill_id, self.user["id"])
        if bill:
            categories = self.category_model.get_categories(self.user["id"], bill[1])
            self.view.show_add_dialog(categories, self.save_bill, bill)

    def handle_delete(self, bill_id):
        success, message = self.bill_model.delete_bill(bill_id, self.user["id"])
        
        if success:
            self.view.show_message(message)
            self.load_bills()
        else:
            self.view.show_message(message, is_error=True)

    def handle_export(self):
        params = self.view.get_filter_params()
        bills = self.bill_model.get_bills(
            self.user["id"],
            type_=params["type"],
            start_date=params["start_date"],
            end_date=params["end_date"]
        )

        if not bills:
            self.view.show_message("没有数据可以导出", is_error=True)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "账单数据"

        headers = ["类型", "日期", "分类", "金额", "付款方式", "备注", "创建时间"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, bill in enumerate(bills, 2):
            bill_id, type_, category_id, category_name, amount, remark, payment_method, date, created_at = bill
            ws.cell(row=row, column=1, value="收入" if type_ == "income" else "支出")
            ws.cell(row=row, column=2, value=date)
            ws.cell(row=row, column=3, value=category_name)
            ws.cell(row=row, column=4, value=amount)
            ws.cell(row=row, column=5, value=payment_method)
            ws.cell(row=row, column=6, value=remark or "")
            ws.cell(row=row, column=7, value=created_at)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"账单导出_{timestamp}.xlsx"
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, filename)

        try:
            wb.save(file_path)
            self.view.show_message(f"导出成功！文件已保存到桌面：{filename}")
        except Exception as e:
            self.view.show_message(f"导出失败：{str(e)}", is_error=True)

    def handle_statistics(self):
        summary = self.bill_model.get_total_summary(self.user["id"])
        monthly_data = self.bill_model.get_monthly_statistics(self.user["id"])
        yearly_data = self.bill_model.get_yearly_statistics(self.user["id"])
        category_data = self.bill_model.get_expense_category_statistics(self.user["id"])

        statistics_data = {
            'total_income': summary.get('income', 0),
            'total_expense': summary.get('expense', 0),
            'monthly_data': monthly_data,
            'yearly_data': yearly_data,
            'category_data': category_data
        }

        self.view.show_statistics_dialog(statistics_data)